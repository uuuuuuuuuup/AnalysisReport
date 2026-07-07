# -*- coding: utf-8 -*-
"""
财务打分模型——输出 sheet 数据提取脚本（v3.0）

【v4.0 状态说明】
自 v4.0 起，本脚本降级为**可选工具**，不再用于常规穿透财报分析报告生成。
v4.0 已删除报告第四部分「全A样本分位数定位」，综合结论改为基于三张表深度分析
的财务质量五维评估。仅当用户**主动要求**进行全A样本对比时，才使用本脚本。

【核心规则——v3.0，v4.0 可选】
1. 全A样本部分的数据（全A中位数规律、个股5大打分指标、分位数、wind一致预期）
   全部来自用户上传的「财务打分模型.xlsx」，不从网络摘取。
2. 脚本只提取「输出」sheet（第一个 sheet）的数据，不遍历年度/季度 sheet。
3. 用户只需在「输出」sheet 的 B4 格填入要分析的公司简称，其余单元格由 Excel 公式
   自动生成（VLOOKUP 等从底层年度/季度 sheet 拉取）。脚本读取这些公式的缓存值。
4. 若用户未上传有效的财务打分模型.xlsx，则提示上传，不生成全A样本对比部分。

【输出 sheet 读取原理】
- 使用 openpyxl data_only=True 读取公式计算后的缓存值。
- 缓存值只有在 Excel/WPS 打开过该文件并保存后才存在。若文件从未被打开，
  公式单元格的值将为 None，脚本会检测到此情况并提示用户用 Excel 打开并保存。
- 脚本会将输出 sheet 的全部非空单元格完整导出（原始矩阵），同时尝试智能
  识别关键结构化字段（公司信息、5大指标最新分位数、年度/季度趋势、wind预期），
  供报告生成器使用。

【公司简称获取方式】
  方式1（推荐，与 Excel 交互一致）：在「输出」sheet 的 B4 格填入公司简称，
          脚本自动读取 B4，无需再传 --company。
  方式2：通过 --company 参数显式指定（优先级高于 B4）。

用法:
    # 方式1：自动读取 B4（用户已在 Excel 的 B4 填好公司简称）
    python financial_scoring.py --model "财务打分模型.xlsx" [--output result.json]

    # 方式2：显式指定公司简称
    python financial_scoring.py --model "财务打分模型.xlsx" --company "宁德时代" [--output result.json]

输出 JSON 格式，包含:
- 公司简称、分析时间、模型路径、数据来源说明
- 输出sheet原始矩阵（所有非空单元格，保留行列位置）
- 智能识别结果：公司基本信息、5大指标最新分位数汇总、年度趋势、季度趋势、wind一致预期
- 缓存值检测状态（是否有公式未计算的单元格）
"""

import argparse
import json
import math
import re
import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl 未安装，请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================
# 关键词定义（用于智能识别输出 sheet 中的数据区块）
# ============================================================

# 5大指标关键词（用于识别指标区块标题行）
INDICATOR_KEYWORDS = {
    "指标1": ["应收款项周转率", "应收款项", "应收周转", "指标1", "指标一"],
    "指标2": ["固定资产", "无形资产", "再投资", "指标2", "指标二"],
    "指标3": ["营运净资本", "营运净资产", "营运资本", "产业链", "指标3", "指标三"],
    "指标4": ["经营性现金流", "经营现金流", "现金流", "指标4", "指标四"],
    "指标5": ["wind", "一致预期", "预测", "指标5", "指标五"],
}

# 数据类型关键词（用于识别行/列的数据类型）
DATA_TYPE_KEYWORDS = {
    "全A中位数": ["中位数", "全a中位", "全a样本中位"],
    "分位数": ["分位", "百分位"],
    "公司值": ["公司", "目标", "本企业"],
    "全A分布": ["分布", "直方", "频次"],
}

# 年份正则（2015-2099）
YEAR_RE = re.compile(r"^(20[0-2]\d|19\d\d)$")
# 季度正则（如 2016Q1）
QUARTER_RE = re.compile(r"^(20[0-2]\d)Q([1-4])$")
# 年度A正则（如 2015A）
YEAR_A_RE = re.compile(r"^(20[0-2]\d)A$")


# ============================================================
# 通用工具函数
# ============================================================

def safe_float(val):
    """安全转换为 float，处理 #N/A、空值、异常字符串"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
            return None
        return float(val)
    s = str(val).strip()
    if s in ("", "#N/A", "#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "None", "nan", "NaN", "null"):
        return None
    # 处理百分号
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100.0
        except (ValueError, TypeError):
            return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def cell_to_float(val):
    """单元格值转 float，与 safe_float 一致"""
    return safe_float(val)


def text_of(val):
    """取单元格文本"""
    if val is None:
        return ""
    return str(val).strip()


def is_year_label(val):
    """是否为年份标签（如 '2015'）"""
    s = text_of(val)
    return bool(YEAR_RE.match(s))


def is_quarter_label(val):
    """是否为季度标签（如 '2016Q1'）"""
    s = text_of(val)
    return bool(QUARTER_RE.match(s))


def is_year_a_label(val):
    """是否为年度A标签（如 '2015A'）"""
    s = text_of(val)
    return bool(YEAR_A_RE.match(s))


def is_time_label(val):
    """是否为任意时间标签（年份/季度/年度A）"""
    return is_year_label(val) or is_quarter_label(val) or is_year_a_label(val)


def match_keywords(text, keyword_groups):
    """检查文本是否包含某组关键词中的任意一个。
    keyword_groups: dict of {group_name: [keywords]}
    返回匹配的 group_name 或 None。
    """
    low = text.lower()
    for group, kws in keyword_groups.items():
        for kw in kws:
            if kw.lower() in low:
                return group
    return None


# ============================================================
# 文件校验
# ============================================================

def validate_model_file(model_path):
    """验证文件是有效的「财务打分模型.xlsx」：
    - 文件存在
    - 可打开
    - 第一个 sheet 名为「输出」
    - B4 格有公司简称（或可通过 --company 指定）
    返回 (是否有效, 错误信息)。
    """
    if not os.path.exists(model_path):
        return False, f"文件不存在: {model_path}"
    try:
        wb = openpyxl.load_workbook(model_path, read_only=True, data_only=True)
    except Exception as e:
        return False, f"无法打开 xlsx 文件: {e}"
    if not wb.sheetnames:
        wb.close()
        return False, "xlsx 文件无任何 sheet"
    first_sheet_name = wb.sheetnames[0]
    if first_sheet_name != "输出":
        wb.close()
        return False, (
            f"第一个 sheet 名称为「{first_sheet_name}」，不是「输出」。"
            f"请确认上传的是「财务打分模型.xlsx」（首个 sheet 应为「输出」，B4 格用于填入公司简称）。"
        )
    wb.close()
    return True, "OK"


def read_company_from_b4(model_path):
    """从「输出」sheet 的 B4 格读取公司简称。
    返回 (公司简称 or None, 错误信息)。
    """
    try:
        wb = openpyxl.load_workbook(model_path, read_only=True, data_only=True)
    except Exception as e:
        return None, f"无法打开 xlsx 文件: {e}"
    ws = wb.worksheets[0]
    b4 = ws["B4"].value
    wb.close()
    if b4 is None or text_of(b4) == "":
        return None, "「输出」sheet 的 B4 格为空，请先在 B4 填入要分析的公司简称。"
    return text_of(b4), "OK"


# ============================================================
# 输出 sheet 全量提取
# ============================================================

def extract_output_sheet_raw(model_path):
    """提取「输出」sheet 的全部单元格数据（data_only=True 读取公式缓存值）。

    返回:
        dict:
            "rows": list of list —— 二维矩阵（按行，每行为各列值），空单元格为 None
            "cells": list of [row_idx(1-based), col_idx(1-based), value] —— 所有非空单元格
            "max_row": int
            "max_col": int
            "has_cached_values": bool —— 是否有数值型缓存值（用于检测公式是否已计算）
            "none_in_data_area": int —— 数据区域中 None 单元格数量（可能为公式未计算）
    """
    wb = openpyxl.load_workbook(model_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    rows = []
    cells = []
    numeric_count = 0
    none_count = 0

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True),
        start=1,
    ):
        row_list = list(row)
        rows.append(row_list)
        for c_idx, val in enumerate(row_list, start=1):
            if val is not None and text_of(val) != "":
                cells.append([r_idx, c_idx, val])
                if isinstance(val, (int, float)) and not (isinstance(val, float) and math.isnan(val)):
                    numeric_count += 1
            else:
                none_count += 1

    wb.close()

    has_cached_values = numeric_count > 5  # 至少有一些数值，说明公式有缓存

    return {
        "rows": rows,
        "cells": cells,
        "max_row": max_row,
        "max_col": max_col,
        "has_cached_values": has_cached_values,
        "numeric_count": numeric_count,
        "none_count": none_count,
    }


# ============================================================
# 智能识别结构化字段
# ============================================================

def build_text_index(cells):
    """构建文本单元格索引，返回:
        - text_cells: list of (row, col, text) 文本类型单元格
        - year_cells: list of (row, col, year_str) 年份标签单元格
        - quarter_cells: list of (row, col, quarter_str) 季度标签单元格
        - year_a_cells: list of (row, col, year_a_str) 年度A标签单元格
    """
    text_cells = []
    year_cells = []
    quarter_cells = []
    year_a_cells = []

    for r, c, val in cells:
        if isinstance(val, str):
            t = val.strip()
            if t == "":
                continue
            text_cells.append((r, c, t))
            if is_year_label(t):
                year_cells.append((r, c, t))
            elif is_quarter_label(t):
                quarter_cells.append((r, c, t))
            elif is_year_a_label(t):
                year_a_cells.append((r, c, t))
    return text_cells, year_cells, quarter_cells, year_a_cells


def find_indicator_blocks(text_cells):
    """识别5大指标区块的起始行。
    返回 dict: {指标序号: [(row, col, text), ...]} 区块标题单元格列表。
    """
    blocks = {}
    for r, c, t in text_cells:
        group = match_keywords(t, INDICATOR_KEYWORDS)
        if group:
            blocks.setdefault(group, []).append((r, c, t))
    return blocks


def extract_time_series_from_row(rows, row_idx, max_col, time_labels_in_row):
    """从某一行提取时间序列数据。
    time_labels_in_row: list of (col, label) 该行中的时间标签及其列位置。
    返回 dict: {label: value}
    """
    result = {}
    if row_idx - 1 < 0 or row_idx - 1 >= len(rows):
        return result
    row_data = rows[row_idx - 1]
    for col, label in time_labels_in_row:
        if col - 1 < len(row_data):
            val = safe_float(row_data[col - 1])
            if val is not None:
                result[label] = val
    return result


def smart_extract(rows, cells, company_name):
    """智能识别输出 sheet 中的结构化字段。

    返回 dict:
        - 公司基本信息
        - 五大指标最新分位数汇总
        - 年度趋势  (各指标的公司值/全A中位数/分位数的年度序列)
        - 季度趋势
        - wind一致预期
        - 识别说明 / 告警
    """
    result = {
        "公司基本信息": {},
        "五大指标最新分位数汇总": {},
        "年度趋势": {},
        "季度趋势": {},
        "wind一致预期": {},
        "识别告警": [],
    }

    text_cells, year_cells, quarter_cells, year_a_cells = build_text_index(cells)

    # ----------------------------------------------------------
    # 1. 识别公司基本信息（代码/行业/上市时间等）
    #    在输出 sheet 中搜索标签行
    # ----------------------------------------------------------
    info_labels = {
        "代码": ["代码", "股票代码"],
        "一级行业": ["一级行业", "行业", "申万"],
        "上市时间": ["上市时间", "上市日期"],
        "上市年份": ["上市年份"],
    }
    for r, c, t in text_cells:
        for field, kws in info_labels.items():
            if any(kw in t for kw in kws) and len(t) < 20:
                # 在同行右侧或下一行取值
                row_data = rows[r - 1] if r - 1 < len(rows) else []
                # 优先取右侧单元格
                val = None
                if c < len(row_data):
                    val = row_data[c]  # 紧邻右侧（同一标签格的下一列）
                if val is None or text_of(val) == "":
                    # 尝试下一行同列
                    if r < len(rows):
                        next_row = rows[r]
                        if c - 1 < len(next_row):
                            val = next_row[c - 1]
                if val is not None and text_of(val) != "":
                    result["公司基本信息"][field] = text_of(val)
                break

    # ----------------------------------------------------------
    # 2. 识别5大指标区块
    # ----------------------------------------------------------
    indicator_blocks = find_indicator_blocks(text_cells)

    # 排序各指标区块的标题行（取每组的最早行）
    block_starts = {}
    for group, items in indicator_blocks.items():
        items_sorted = sorted(items, key=lambda x: x[0])
        block_starts[group] = items_sorted[0][0]  # 起始行

    # 按行号排序区块，确定区块边界
    sorted_blocks = sorted(block_starts.items(), key=lambda x: x[1])

    for i, (group, start_row) in enumerate(sorted_blocks):
        end_row = sorted_blocks[i + 1][1] if i + 1 < len(sorted_blocks) else len(rows) + 1
        block_rows = [r for r in range(start_row, min(end_row, len(rows) + 1))]

        # 在该区块内找时间标签行（年度/季度）
        block_year_labels = []  # [(row, [(col, label), ...])]
        block_quarter_labels = []
        for r in block_rows:
            if r - 1 >= len(rows):
                continue
            row_data = rows[r - 1]
            yr_in_row = []
            qs_in_row = []
            for c_idx, val in enumerate(row_data, start=1):
                if is_year_label(val):
                    yr_in_row.append((c_idx, text_of(val)))
                elif is_quarter_label(val):
                    qs_in_row.append((c_idx, text_of(val)))
            if yr_in_row:
                block_year_labels.append((r, yr_in_row))
            if qs_in_row:
                block_quarter_labels.append((r, qs_in_row))

        # 在该区块内找数据类型标签（中位数/分位/公司值）
        block_data_types = []  # [(row, col, type_name, text)]
        for r in block_rows:
            if r - 1 >= len(rows):
                continue
            for c_idx, val in enumerate(rows[r - 1], start=1):
                if isinstance(val, str):
                    dtype = match_keywords(val, DATA_TYPE_KEYWORDS)
                    if dtype:
                        block_data_types.append((r, c_idx, dtype, val.strip()))

        # 提取年度趋势：找到有年份标签的行，提取该行各年份对应的数据
        # 也找"中位数"/"分位"/"公司"行，提取其对应年份列的数据
        if group not in result["年度趋势"]:
            result["年度趋势"][group] = {}
        if group not in result["季度趋势"]:
            result["季度趋势"][group] = {}

        # 策略：找到列头含年份的行，确定"年份->列"映射，然后向下找各数据行
        year_col_map = {}  # {year_label: col}
        for r, yr_list in block_year_labels:
            for col, label in yr_list:
                year_col_map[label] = col

        quarter_col_map = {}
        for r, qs_list in block_quarter_labels:
            for col, label in qs_list:
                quarter_col_map[label] = col

        # 对每个数据类型行，提取年度序列
        for r, c, dtype, text in block_data_types:
            row_data = rows[r - 1] if r - 1 < len(rows) else []
            year_series = {}
            for yr_label, col in year_col_map.items():
                if col - 1 < len(row_data):
                    v = safe_float(row_data[col - 1])
                    if v is not None:
                        year_series[yr_label] = v
            quarter_series = {}
            for qs_label, col in quarter_col_map.items():
                if col - 1 < len(row_data):
                    v = safe_float(row_data[col - 1])
                    if v is not None:
                        quarter_series[qs_label] = v
            if year_series:
                result["年度趋势"][group].setdefault(dtype, {}).setdefault(
                    f"行{r}_{text[:10]}", {}
                ).update(year_series)
            if quarter_series:
                result["季度趋势"][group].setdefault(dtype, {}).setdefault(
                    f"行{r}_{text[:10]}", {}
                ).update(quarter_series)

        # 如果没有找到数据类型标签行，直接从年份标签行提取数据
        if not block_data_types and year_col_map:
            for r, yr_list in block_year_labels:
                row_data = rows[r - 1] if r - 1 < len(rows) else []
                series = {}
                for col, label in yr_list:
                    # 年份标签格本身的值是文本，取下一列或同行其他列的数据
                    if col < len(row_data):
                        v = safe_float(row_data[col])  # 年份右侧一格
                        if v is not None:
                            series[label] = v
                if series:
                    result["年度趋势"][group].setdefault("数据行", {}).setdefault(
                        f"行{r}", {}
                    ).update(series)

    # ----------------------------------------------------------
    # 3. 提取 wind 一致预期（指标5区块）
    # ----------------------------------------------------------
    if "指标5" in block_starts:
        start = block_starts["指标5"]
        end = len(rows) + 1
        for i, (g, s) in enumerate(sorted_blocks):
            if g == "指标5" and i + 1 < len(sorted_blocks):
                end = sorted_blocks[i + 1][1]
        # 在指标5区块内，找所有时间标签（含 Q1-Q4 和 A）
        wind_data = {}
        for r in range(start, min(end, len(rows) + 1)):
            if r - 1 >= len(rows):
                continue
            row_data = rows[r - 1]
            for c_idx, val in enumerate(row_data, start=1):
                t = text_of(val)
                if is_quarter_label(t) or is_year_a_label(t) or is_year_label(t):
                    # 取该标签右侧的数据值
                    if c_idx < len(row_data):
                        v = safe_float(row_data[c_idx])
                        if v is not None:
                            wind_data[t] = v
        result["wind一致预期"] = wind_data

    # ----------------------------------------------------------
    # 4. 提取最新分位数汇总
    #    从各指标区块的季度趋势中取最新季度的分位数
    # ----------------------------------------------------------
    indicator_pct_map = {
        "指标1": "指标1_应收款项周转率分位数",
        "指标2": "指标2_固定资产周转率分位数",
        "指标3": "指标3_营运资本占比分位数",
        "指标4": "指标4_经营现金流分位数",
    }
    for group, label in indicator_pct_map.items():
        # 从年度趋势中取最新年份的分位数
        if group in result["年度趋势"]:
            for dtype, rows_dict in result["年度趋势"][group].items():
                if "分位" in dtype:
                    for row_key, series in rows_dict.items():
                        if series:
                            latest_year = max(series.keys())
                            val = series[latest_year]
                            pct = round(val * 100, 1) if val <= 1 else round(val, 1)
                            result["五大指标最新分位数汇总"][f"{latest_year}年报_{label}"] = pct
                            break
                    break
        # 从季度趋势中取最新季度的分位数
        if group in result["季度趋势"]:
            for dtype, rows_dict in result["季度趋势"][group].items():
                if "分位" in dtype:
                    for row_key, series in rows_dict.items():
                        if series:
                            # 最新季度（按时间排序）
                            def sort_key(k):
                                m = QUARTER_RE.match(k)
                                if m:
                                    return (int(m.group(1)), int(m.group(2)))
                                return (0, 0)
                            latest_q = max(series.keys(), key=sort_key)
                            val = series[latest_q]
                            pct = round(val * 100, 1) if val <= 1 else round(val, 1)
                            result["五大指标最新分位数汇总"][f"{latest_q}_{label}"] = pct
                            break
                    break

    # ----------------------------------------------------------
    # 5. 告警信息
    # ----------------------------------------------------------
    if not result["年度趋势"]:
        result["识别告警"].append(
            "未能从输出 sheet 识别出年度趋势数据。可能原因：输出 sheet 公式未计算（请用 Excel 打开文件并保存），或布局与预期不符。"
        )
    if not result["五大指标最新分位数汇总"]:
        result["识别告警"].append(
            "未能提取5大指标最新分位数。请检查输出 sheet 是否已正确生成分位数数据。"
        )

    return result


# ============================================================
# 主分析函数
# ============================================================

def analyze_company(model_path, company_name):
    """主分析函数：只从「输出」sheet 提取目标公司的全部指标。"""
    print(f"加载财务打分模型: {model_path}")
    print(f"分析公司: {company_name}")
    print(f"数据来源: 用户上传的财务打分模型.xlsx「输出」sheet（全A样本数据均来自此文件，不从网络摘取）")

    # 提取输出 sheet 原始数据
    print("\n--- 提取「输出」sheet 数据 ---")
    raw = extract_output_sheet_raw(model_path)
    print(f"  输出 sheet 尺寸: {raw['max_row']} 行 × {raw['max_col']} 列")
    print(f"  非空单元格: {len(raw['cells'])} 个")
    print(f"  数值型单元格: {raw['numeric_count']} 个")

    if not raw["has_cached_values"]:
        print("\n  ⚠ 警告: 输出 sheet 中数值型单元格很少（{}个）。".format(raw['numeric_count']))
        print("  这通常意味着 Excel 公式尚未计算（文件未被 Excel/WPS 打开并保存）。")
        print("  请用 Excel/WPS 打开「财务打分模型.xlsx」，确认「输出」sheet 数据正常显示后保存，再重新运行。")

    # 智能识别结构化字段
    print("\n--- 智能识别结构化字段 ---")
    smart = smart_extract(raw["rows"], raw["cells"], company_name)

    company_info = smart.get("公司基本信息", {})
    if company_info:
        print(f"  公司基本信息: 代码={company_info.get('代码')}, 行业={company_info.get('一级行业')}")
    else:
        print("  公司基本信息: 未识别到（可从输出 sheet 手动确认）")

    pct_summary = smart.get("五大指标最新分位数汇总", {})
    if pct_summary:
        print(f"  5大指标最新分位数汇总: {len(pct_summary)} 项")
        for k, v in pct_summary.items():
            print(f"    {k}: {v}%")
    else:
        print("  5大指标最新分位数汇总: 未识别到")

    for alert in smart.get("识别告警", []):
        print(f"  [告警] {alert}")

    # 组装结果
    result = {
        "公司简称": company_name,
        "分析时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "模型路径": os.path.abspath(model_path),
        "数据来源说明": (
            "全部全A样本数据（中位数规律、个股5大打分指标、分位数、wind一致预期）均来自"
            "用户上传的财务打分模型.xlsx的「输出」sheet，未从网络摘取。"
            "用户在B4格填入公司简称后，输出sheet由Excel公式自动生成，脚本只提取输出sheet的数据。"
        ),
        "缓存值检测": {
            "有数值缓存": raw["has_cached_values"],
            "数值型单元格数": raw["numeric_count"],
            "非空单元格数": len(raw["cells"]),
        },
        "公司基本信息": company_info,
        "五大指标最新分位数汇总": pct_summary,
        "年度趋势": smart.get("年度趋势", {}),
        "季度趋势": smart.get("季度趋势", {}),
        "wind一致预期": smart.get("wind一致预期", {}),
        "识别告警": smart.get("识别告警", []),
        "输出sheet_尺寸": {
            "行数": raw["max_row"],
            "列数": raw["max_col"],
        },
        "输出sheet_原始矩阵": raw["rows"],
        "输出sheet_非空单元格": raw["cells"],
    }

    return result


# ============================================================
# 命令行入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="全A样本财务打分模型分析（v3.0——只从输出sheet提取数据，强制用户上传）"
    )
    parser.add_argument("--model", required=True,
                        help="【必填】用户上传的「财务打分模型.xlsx」路径")
    parser.add_argument("--company", default=None,
                        help="公司简称（可选；若未提供，则自动从「输出」sheet 的 B4 格读取）")
    parser.add_argument("--output", default=None,
                        help="输出 JSON 文件路径（可选，默认为 {公司简称}_财务打分结果.json）")
    args = parser.parse_args()

    # 1. 校验模型文件
    ok, msg = validate_model_file(args.model)
    if not ok:
        print(f"ERROR: {msg}", file=sys.stderr)
        print("提示：请上传有效的「财务打分模型.xlsx」（首个 sheet 应为「输出」，"
              "在 B4 格填入要分析的公司简称）。未上传有效文件将不生成全A样本对比部分。",
              file=sys.stderr)
        sys.exit(1)

    # 2. 确定公司简称
    company = args.company
    if company:
        company = company.strip()
        print(f"使用 --company 参数指定的公司简称: {company}")
    else:
        company, msg = read_company_from_b4(args.model)
        if company is None:
            print(f"ERROR: {msg}", file=sys.stderr)
            print("提示：请在「输出」sheet 的 B4 格填入要分析的公司简称，"
                  "或通过 --company 参数指定。", file=sys.stderr)
            sys.exit(1)
        print(f"自动从「输出」sheet 的 B4 格读取公司简称: {company}")

    # 3. 执行分析
    result = analyze_company(args.model, company)

    # 4. 输出 JSON
    output_path = args.output or f"{company}_财务打分结果.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)

    print(f"\n结果已保存到: {output_path}")
    print(f"公司: {company}")
    if result.get("公司基本信息"):
        print(f"代码: {result['公司基本信息'].get('代码')}")
        print(f"行业: {result['公司基本信息'].get('一级行业')}")
    print(f"年度趋势指标组: {len(result.get('年度趋势', {}))} 个")
    print(f"季度趋势指标组: {len(result.get('季度趋势', {}))} 个")
    print(f"wind 预期: {len(result.get('wind一致预期', {}))} 个标签")
    print(f"\n五大指标最新分位数汇总:")
    for k, v in result.get("五大指标最新分位数汇总", {}).items():
        print(f"  {k}: {v}%")
    print(f"\n数据来源: 用户上传的财务打分模型.xlsx「输出」sheet（全A样本数据均来自此文件，未从网络摘取）")
    print(f"输出sheet原始矩阵: {result['输出sheet_尺寸']['行数']} 行 × {result['输出sheet_尺寸']['列数']} 列")


if __name__ == "__main__":
    main()
